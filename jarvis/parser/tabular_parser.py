"""
parser/tabular_parser.py
=========================

CSV and Excel (.xls/.xlsx) parsing.

Design decisions
-----------------
- Uses the stdlib `csv` module (not pandas) for CSV — pandas is available
  in this environment but is a heavy dependency to *require* on a
  CPU-only, install-restricted Citrix box just to read delimited text, so
  we avoid depending on it for the common case.
- Uses `openpyxl` for `.xlsx` (already required). For legacy `.xls`
  (pre-2007 binary format) we attempt `openpyxl` first (which cannot read
  true `.xls`) and fail clearly with a actionable message rather than
  silently returning nothing — `.xls` support requires `xlrd` which is a
  separate optional dependency the deployment can add if needed.
- Each row is NOT emitted as a separate element (that would explode chunk
  count on a 10,000-row BOM). Instead, a whole sheet/CSV is rendered as one
  or more TABLE elements in a compact pipe-delimited text form that stays
  both human-readable and grep/keyword-searchable, batched at a bounded
  row count per element (`_ROWS_PER_CHUNK`) so a huge spreadsheet still
  chunks reasonably for retrieval instead of becoming one giant blob.
- Column headers are repeated in every batch's rendered text so a search
  hit deep in a large table still shows which column is which.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import List

from models.document import ExtractionElementType
from parser.base_parser import BaseParser, ParsedDocument, ParserError, RawElement
from utils.logger import get_logger

logger = get_logger(__name__)

_ROWS_PER_CHUNK = 50


def _render_rows(headers: List[str], rows: List[List[str]]) -> str:
    lines = [" | ".join(headers)] if headers else []
    lines += [" | ".join(str(c) for c in row) for row in rows]
    return "\n".join(lines)


class CsvParser(BaseParser):
    @classmethod
    def supported_extensions(cls) -> List[str]:
        return [".csv"]

    def parse(self, path: Path) -> ParsedDocument:
        from parser.text_family_parser import _read_text_with_encoding_detection

        text = _read_text_with_encoding_detection(path)
        doc = ParsedDocument(title=path.stem, page_count=1)

        try:
            dialect = csv.Sniffer().sniff(text[:4096])
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(text.splitlines(), dialect)
        rows = list(reader)
        if not rows:
            return doc

        headers, data_rows = rows[0], rows[1:]
        order = 0
        for i in range(0, len(data_rows), _ROWS_PER_CHUNK):
            batch = data_rows[i:i + _ROWS_PER_CHUNK]
            doc.elements.append(RawElement(
                text=_render_rows(headers, batch),
                element_type=ExtractionElementType.TABLE,
                section_path=f"rows {i + 1}-{i + len(batch)}",
                order_index=order,
                extra={"headers": headers, "row_count": len(batch)},
            ))
            order += 1
        return doc


class ExcelParser(BaseParser):
    @classmethod
    def supported_extensions(cls) -> List[str]:
        return [".xlsx", ".xls"]

    def parse(self, path: Path) -> ParsedDocument:
        import openpyxl

        try:
            workbook = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        except Exception as exc:  # openpyxl raises several distinct exception types
            raise ParserError(f"Could not open workbook {path}: {exc}")

        doc = ParsedDocument(title=path.stem, page_count=len(workbook.sheetnames))
        order = 0

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            all_rows = []
            for row in sheet.iter_rows(values_only=True):
                if row is None or all(cell is None for cell in row):
                    continue
                all_rows.append(["" if c is None else c for c in row])

            if not all_rows:
                continue

            headers = [str(h) for h in all_rows[0]]
            data_rows = all_rows[1:]
            if not data_rows:
                data_rows = [all_rows[0]]
                headers = [f"col_{i+1}" for i in range(len(all_rows[0]))]

            for i in range(0, len(data_rows), _ROWS_PER_CHUNK):
                batch = data_rows[i:i + _ROWS_PER_CHUNK]
                doc.elements.append(RawElement(
                    text=_render_rows(headers, batch),
                    element_type=ExtractionElementType.TABLE,
                    section_path=f"{sheet_name} (rows {i + 1}-{i + len(batch)})",
                    order_index=order,
                    extra={"sheet": sheet_name, "headers": headers, "row_count": len(batch)},
                ))
                order += 1

        workbook.close()
        return doc
